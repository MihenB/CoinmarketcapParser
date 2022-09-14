import datetime
from parse_package.multypurpose_parser import ScrapSession
from config import url, headers, params
import openpyxl

wb = None
sheet = None
titles = (
    'name',
    'rank',
    'price',
    'ath',
    'atl'
)


def create_excel_sheet():
    global wb, sheet
    wb = openpyxl.load_workbook(filename='coinmarketcap.xlsx')
    sheet = wb['Sheet1']
    row = 1
    for pos, title in enumerate(titles, start=1):
        sheet.cell(row=row, column=pos).value = title


def write_to_excel(vals, row):
    global sheet
    for i, rec in enumerate(vals, start=1):
        sheet.cell(row=row, column=i).value = rec


def get_last_item(session):
    return int(session.get(
        url=url,
        headers=headers,
        params=params
    ).json['data']['totalCount'])


def get_data():
    global wb
    start_time = datetime.datetime.now()
    create_excel_sheet()
    session = ScrapSession()
    row = 1
    for i in range(1, get_last_item(session), int(params['limit'])):
        params.update(start=i)
        json = session.get(
            url=url,
            headers=headers,
            params=params
        ).json['data']['cryptoCurrencyList']
        for item in json:
            row += 1
            name = item['name']
            rank = item['cmcRank']
            price = item['quotes'][2]['price']  # USD
            ath = item['ath']
            atl = item['atl']
            write_to_excel(
                (name, rank, price, ath, atl),
                row
            )
    wb.save('coinmarketcap.xlsx')
    end_time = datetime.datetime.now() - start_time
    print('Execution time: ', end_time)


def main():
    get_data()


if __name__ == '__main__':
    main()
