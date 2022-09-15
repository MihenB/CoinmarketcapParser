import asyncio
import datetime
import json
from config import url, headers, params
import openpyxl
from aiohttp import ClientSession


def create_excel_sheet(titles: tuple):
    _wb = openpyxl.Workbook()
    _sheet = _wb['Sheet']
    row = 1
    for pos, title in enumerate(titles, start=1):
        _sheet.cell(row=row, column=pos).value = title
    return _wb, _sheet


async def get_last_item(session: ClientSession):
    response = await session.get(
        url=url,
        headers=headers,
        params=params)
    return int(json.loads(await response.text())['data']['totalCount'])


async def request_to_data(start: int, session: ClientSession, _sheet):
    params_copy = params.copy()
    params_copy.update(start=start)
    response = await session.get(
        url=url,
        headers=headers,
        params=params_copy
    )
    json_ = json.loads(await response.text())['data']['cryptoCurrencyList']
    for item in json_:
        name = item['name']
        rank = item['cmcRank']
        price = item['quotes'][2]['price']  # USD
        ath = item['ath']
        atl = item['atl']
        coefficient = get_coefficient(rank, price, ath, atl)
        write_to_excel(
            (name, coefficient),
            rank + 1,
            _sheet
        )


def get_coefficient(rank, price, ath, atl):
    return (ath * atl / price ** 2) / rank if price != 0 else 'PRICE EQUALS ZERO'


def write_to_excel(vals, row, _sheet):
    for i, rec in enumerate(vals, start=1):
        _sheet.cell(row=row, column=i).value = rec


def exec_time_decorator(func):
    async def wrapper(*args, **kwargs):
        start_time = datetime.datetime.now()
        result = await func(*args, **kwargs)
        print(datetime.datetime.now() - start_time)
        return result

    return wrapper


def format_col_width(sheet):
    import string
    for letter in string.ascii_uppercase[:2]:
        sheet.column_dimensions[letter].width = 40


@exec_time_decorator
async def get_data() -> str:
    titles = (
        'name',
        'coefficient'
    )

    file_name = f'{datetime.datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}_coinmarketcap.xlsx'

    wb, ws = create_excel_sheet(titles=titles)

    async with ClientSession() as session:
        last_item_num = await get_last_item(session)
        step = int(params['limit'])
        tasks = [
            asyncio.create_task(request_to_data(i, session, ws)) for i in range(1, last_item_num, step)
        ]
        await asyncio.gather(*tasks)

    ws.auto_filter.ref = f'A1:B{last_item_num + 1}'
    ws.auto_filter.add_sort_condition(f'B2:B{last_item_num + 1}')

    format_col_width(sheet=ws)

    wb.save(filename=file_name)

    return file_name


async def main():
    print(await get_data())


if __name__ == '__main__':
    asyncio.run(main())
