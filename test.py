import asyncio
import datetime
import json
from config import url, headers, params
import openpyxl
from aiohttp import ClientSession


wb = None
sheet = None
titles = (
    'name',
    'rank',
    'price',
    'ath',
    'atl'
)


def create_excel_sheet():
    global wb, sheet
    wb = openpyxl.load_workbook(filename='coinmarketcap.xlsx')
    sheet = wb['Sheet1']
    row = 1
    for pos, title in enumerate(titles, start=1):
        sheet.cell(row=row, column=pos).value = title


def write_to_excel(vals, row):
    global sheet
    for i, rec in enumerate(vals, start=1):
        sheet.cell(row=row, column=i).value = rec


async def get_last_item(session: ClientSession):
    response = await session.get(
        url=url,
        headers=headers,
        params=params)
    return int(json.loads(await response.text())['data']['totalCount'])


async def request_to_data(start: int, session: ClientSession):
    params_copy = params.copy()
    params_copy.update(start=start)
    response = await session.get(
        url=url,
        headers=headers,
        params=params_copy
    )
    json_ = json.loads(await response.text())['data']['cryptoCurrencyList']
    for item in json_:
        name = item['name']
        rank = item['cmcRank']
        price = item['quotes'][2]['price']  # USD
        ath = item['ath']
        atl = item['atl']
        write_to_excel(
            (name, rank, price, ath, atl),
            rank + 1
        )


async def get_data():
    global wb
    start_time = datetime.datetime.now()
    create_excel_sheet()
    async with ClientSession() as session:
        last_item_num = await get_last_item(session)
        print(await request_to_data(1, session))
        tasks = [
            asyncio.create_task(request_to_data(i, session)) for i in range(1, last_item_num, int(params['limit']))
        ]
        await asyncio.gather(*tasks)
    wb.save('coinmarketcap.xlsx')
    time = datetime.datetime.now() - start_time
    print('Execution time: ', time)


async def main():
    await get_data()


if __name__ == '__main__':
    asyncio.run(main())
